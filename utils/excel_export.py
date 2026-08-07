"""
excel_export.py

Builds a formatted .xlsx workbook from a match run's results:
    - "Matches" sheet: one row per YP<->MCP match, enriched with full YP/MCP
      details (name, trade area, address, landmark, phone) looked up from
      the original uploaded records — the match result itself only carries
      IDs, so build_results_workbook() needs the original yps/mcps lists to
      join against.
    - "Waitlist" sheet: one row per unmatched YP + reason, similarly enriched.
    - "Summary" sheet: counts by round, so it's obvious at a glance how many
      matches came from the same-landmark pass vs. each fallback hop.

No formulas here — this is a static results export, not a model — so nothing
needs recalculation before it's handed to the user.

ASSUMPTIONS (confirm these match your actual Person/MCP model from
utils/data_loader.py — adjust the attribute names below if not):
    - YP/MCP objects expose `.name` and `.phone_number`
    - "Trade area" = the existing `.skill` attribute
    - STATUS = whether a match came from the home landmark (round 1) or a
      fallback hop (round 2+), matching the language already used on the
      Summary sheet
    - WAITLISTED column is left blank intentionally, for staff to mark once
      a waitlisted YP is manually resolved — every row on that sheet is
      waitlisted by definition, so a pre-filled value there would be inert
"""

import logging
import time
from collections import defaultdict
from datetime import datetime, timezone
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from configs.schemas import MatchRunResponse

logger = logging.getLogger(__name__)

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
BODY_FONT = Font(name="Arial")
TITLE_FONT = Font(name="Arial", bold=True, size=14)


def _write_table(ws, headers: list[str], rows: list[list], start_row: int = 1):
    logger.debug(
        "_write_table() writing sheet=%r: %d header(s), %d row(s), start_row=%d",
        ws.title, len(headers), len(rows), start_row,
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(rows, start=start_row + 1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = BODY_FONT

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # auto-width, capped so one long address doesn't blow out the sheet
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row in rows:
            val = row[col_idx - 1]
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        width = min(max_len + 3, 45)
        ws.column_dimensions[col_letter].width = width
        if max_len + 3 > 45:
            logger.debug(
                "_write_table() sheet=%r column %r truncated width to cap (content wanted %d)",
                ws.title, header, max_len + 3,
            )

    logger.debug("_write_table() finished sheet=%r", ws.title)


def _build_lookup(people: Iterable) -> dict:
    """id -> person object, for joining match/waitlist rows back to full records."""
    lookup = {}
    for p in people or []:
        lookup[p.id] = p
    return lookup


def _compute_landmark_centroids(yps: Iterable, mcps: Iterable) -> dict:
    """
    landmark name -> "lat, lon" string, averaged over every YP + MCP at that
    landmark with known coordinates. Landmarks with no geocoded people yield
    no centroid (blank in the sheet) rather than a misleading (0, 0).
    """
    points_by_landmark = defaultdict(list)
    for p in list(yps or []) + list(mcps or []):
        if getattr(p, "latitude", None) is not None and getattr(p, "longitude", None) is not None:
            points_by_landmark[p.landmark].append((p.latitude, p.longitude))

    centroids = {}
    for landmark, points in points_by_landmark.items():
        avg_lat = sum(pt[0] for pt in points) / len(points)
        avg_lon = sum(pt[1] for pt in points) / len(points)
        centroids[landmark] = f"{avg_lat:.5f}, {avg_lon:.5f}"

    logger.debug("_compute_landmark_centroids() computed centroids for %d landmark(s)", len(centroids))
    return centroids


def _status_for_round(round_number: int) -> str:
    if round_number == 1:
        return "Matched (Same Landmark)"
    return f"Matched (Fallback Hop {round_number - 1})"


def _field(obj, attr: str, default=""):
    """Best-effort attribute access so a missing field degrades to a blank
    cell instead of crashing the whole export."""
    if obj is None:
        return default
    val = getattr(obj, attr, default)
    return val if val is not None else default


def build_results_workbook(
    result: MatchRunResponse,
    yps: Optional[Iterable] = None,
    mcps: Optional[Iterable] = None,
    output_path: str = "match_results.xlsx",
) -> str:
    """
    Write `result` (matches + waitlist) to output_path as a formatted .xlsx.

    `yps` / `mcps` should be the SAME lists passed into Matcher.run() for
    this result — they're used to look up full YP/MCP details (name, trade
    area, address, landmark, phone) by ID, since the match result itself
    only carries IDs. If omitted, those columns are left blank rather than
    the export failing outright.

    Returns output_path for convenience.
    """
    logger.info(
        "build_results_workbook() starting: %d match(es), %d waitlisted, output_path=%r",
        len(result.matches), len(result.waitlist), output_path,
    )

    start = time.monotonic()

    yp_lookup = _build_lookup(yps)
    mcp_lookup = _build_lookup(mcps)
    landmark_centroids = _compute_landmark_centroids(yps, mcps)

    if yps is not None and len(yp_lookup) < len(result.matches):
        logger.debug(
            "build_results_workbook() yp_lookup has %d entries for %d matches — "
            "some rows may have missing YP details if IDs don't line up",
            len(yp_lookup), len(result.matches),
        )

    wb = Workbook()

    # -- Matches sheet -----------------------------------------------------
    ws_matches = wb.active
    ws_matches.title = "Matches"
    match_headers = [
        "SN", "YP ID YES", "YP NAME", "YP TRADE AREA", "YP ADDRESS", "YP LANDMARK",
        "YP PHONE NUMBER", "MCP ID", "MCP NAME", "MCP ADDRESS", "MCP LANDMARK",
        "MCP TRADE AREA", "LANDMARK CENTROID", "TRAVEL TIME (MINS)", "MATCH ROUND", "STATUS",
    ]

    match_rows = []
    missing_yp_lookups = 0
    missing_mcp_lookups = 0

    for sn, m in enumerate(result.matches, start=1):
        yp = yp_lookup.get(m.yp_id)
        mcp = mcp_lookup.get(m.mcp_id)

        if yps is not None and yp is None:
            missing_yp_lookups += 1
        if mcps is not None and mcp is None:
            missing_mcp_lookups += 1

        match_rows.append([
            sn,
            m.yp_id,
            _field(yp, "name"),
            _field(yp, "skill"),
            _field(yp, "address"),
            _field(yp, "landmark"),
            _field(yp, "phone_number"),
            m.mcp_id,
            _field(mcp, "name"),
            _field(mcp, "address"),
            _field(mcp, "landmark"),
            _field(mcp, "skill"),
            landmark_centroids.get(m.landmark, ""),
            round(m.travel_time, 1),
            m.round,
            _status_for_round(m.round),
        ])

    if missing_yp_lookups or missing_mcp_lookups:
        logger.warning(
            "build_results_workbook() could not find %d YP(s) and %d MCP(s) in the "
            "provided lookup lists — those rows have blank detail columns. Make sure "
            "the yps/mcps passed in are the same ones used for this match run.",
            missing_yp_lookups, missing_mcp_lookups,
        )

    if not match_rows:
        logger.warning("build_results_workbook() no matches to write — Matches sheet will be header-only")

    _write_table(ws_matches, match_headers, match_rows)

    # -- Waitlist sheet -----------------------------------------------------
    ws_waitlist = wb.create_sheet("Waitlist")
    waitlist_headers = ["YP ID", "YP NAME", "PHONE NUMBER", "REASON", "LANDMARK", "WAITLISTED"]

    waitlist_rows = []
    for w in result.waitlist:
        yp = yp_lookup.get(w.yp_id)
        waitlist_rows.append([
            w.yp_id,
            _field(yp, "name"),
            _field(yp, "phone_number"),
            w.reason,
            _field(yp, "landmark"),
            "",  # left blank for staff to mark once manually resolved
        ])

    _write_table(ws_waitlist, waitlist_headers, waitlist_rows)

    # -- Summary sheet -----------------------------------------------------
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "YP <-> MCP Matching Results"
    ws_summary["A1"].font = TITLE_FONT
    ws_summary["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws_summary["A2"].font = BODY_FONT

    rounds_seen = sorted(set(m.round for m in result.matches))
    logger.debug("build_results_workbook() rounds present in results: %s", rounds_seen)

    summary_rows = [["Total matched", result.matched_count]]
    summary_rows += [
        [f"Matched in round {r}" + (" (same landmark)" if r == 1 else " (fallback hop)"),
         sum(1 for m in result.matches if m.round == r)]
        for r in rounds_seen
    ]
    summary_rows.append(["Waitlisted", result.waitlisted_count])

    _write_table(ws_summary, ["Metric", "Count"], summary_rows, start_row=4)

    try:
        wb.save(output_path)
    except OSError:
        logger.exception("build_results_workbook() failed to save workbook to %r", output_path)
        raise

    elapsed = time.monotonic() - start
    logger.info("build_results_workbook() done in %.2fs: saved to %r", elapsed, output_path)

    return output_path
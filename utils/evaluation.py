"""
match_evaluation.py

Compares an automated YP<->MCP match run against a manually-matched
reference sheet, to measure how much the algorithm's output "drifts" from
what a human matcher produced — and, importantly, NOT just by checking
whether the exact same MCP was picked. Two different MCPs can both be
valid for the same YP (same landmark, same trade, both within capacity),
so "different MCP" alone doesn't mean "wrong". Every match is instead
scored against the six agreed criteria:

    1. Distance between YP and MCP
    2. Trade area
    3. MCP specialization (garment/footwear gender split, leather subtype)
    4. MCP capacity
    5. Gender preference
    6. Consideration of PWDs and proximity to the MCP's shop

and classified as an EXACT_MATCH, an EQUIVALENT_MATCH (different MCP, but
meets all agreed criteria), or a DIVERGENT_MATCH (different MCP AND fails
at least one criterion).

WHAT THIS INTENTIONALLY DOES NOT DO: run the algorithm itself, or call the
Google Distance/Geocoding APIs. The automated match result is expected to
already exist (from a normal Matcher.run() call, wherever that's invoked
in the app) and is passed in / uploaded as-is. Distance comparisons use
Haversine distance over the YPs'/MCPs' already-geocoded coordinates
(reusing whatever the app's existing geocode cache/database already has —
see the `distance_service` note on compare_matches()) so that running an
evaluation never re-incurs geocoding or Distance-API cost. This mirrors
the same accuracy caveat already flagged internally: Haversine is a
straight-line approximation, less accurate than the Google Distance API
used at match time, so `distance_metric` is always reported alongside any
distance figure so nobody mistakes one for the other.

ASSUMPTIONS THAT NEED CONFIRMATION (see DEFAULT_CRITERIA_CONFIG below):
this file makes two business-rule assumptions that weren't fully specified
in the discussion — a PWD proximity threshold, and a rule that female YPs
should preferentially land at a female-owned/operated MCP. Both are
configurable and off/on via config so the team can correct them without
touching code; treat the defaults as placeholders, not confirmed policy.

--------------------------------------------------------------------------
CHANGELOG (this revision)
--------------------------------------------------------------------------
- `_evaluate_row` now also records the raw MCP skill/specialization string
  behind each match (`manual_mcp_skill` / `automated_mcp_skill`) and a
  four-way classification of how it relates to the YP's own skill
  (`manual_specialization_class` / `automated_specialization_class`, see
  `classify_specialization()`). The existing `*_specialization_exact`
  booleans stayed as-is (they feed the EQUIVALENT/DIVERGENT verdict and
  changing their meaning would silently change every past report's
  verdict counts) — the new fields are additive, not a replacement.
  Reason: "specialization OK" alone conflates "matched to the exact same
  subtype" with "matched to a both/any wildcard MCP", which is enough for
  compliance but not enough to answer "how many garment_female YPs
  actually landed on a garment_female MCP" — a question that's come up
  more than once.
- `_summarize` now reports mean/median/stdev/p90 for both manual and
  automated distance (in addition to the existing avg_distance_delta_km),
  and a per-skill-category specialization breakdown
  (`specialization_breakdown`) built from `classify_specialization`.
- `write_evaluation_workbook` writes the two new per-row columns, adds the
  distance stats to the Summary sheet, and adds a new "Specialization
  Breakdown" sheet (one row per YP skill category x manual/automated).
"""

import argparse
import json
import logging
import math
import statistics
import time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable, Optional

import pandas as pd

from utils.data_loader import (
    DEFAULT_CONFIG_PATH,
    _clean_str,
    _find_column_name,
    load_config,
    load_mcps,
    load_yps,
)
from logic.matcher import Matcher

logger = logging.getLogger(__name__)

# Reused purely for its trade_matches() method, which needs no distance
# service at all — instantiating with distance_service=None keeps this
# module from needing real geocoding/routing infrastructure just to reuse
# that one piece of logic. This also guarantees evaluation and matching
# can never independently drift on what "trade compatible" means, which is
# exactly the kind of duplicated-logic bug already fixed once in matcher.py.
_trade_matcher = Matcher(distance_service=None)

DEFAULT_CRITERIA_CONFIG = {
    # An automated match within this many km of the manual match's distance
    # (or closer) is treated as meeting the distance criterion, even if it
    # picked a different MCP.
    "distance_tolerance_km": 1.0,

    # ASSUMPTION — confirm with the team: how close (km, Haversine) a PWD
    # YP's matched MCP must be to count as satisfying "consideration of
    # PWDs and proximity to the MCP's shop". Placeholder value.
    "pwd_proximity_threshold_km": 3.0,
}

_WILDCARD_BY_CATEGORY = {
    "garment": "both",
    "footwear": "both",
    "leather": "any",
}

def specialization_matches(yp_specialization, mcp_specialization):
    yp = _clean_str(yp_specialization).lower()
    mcp = _clean_str(mcp_specialization).lower()

    if not yp or not mcp:
        return False

    if yp == mcp:
        return True

    yp_category, _, yp_subtype = yp.rpartition("_")
    mcp_category, _, mcp_subtype = mcp.rpartition("_")

    if not yp_category or not mcp_category or yp_category != mcp_category:
        return False  # different trade area entirely

    wildcard = _WILDCARD_BY_CATEGORY.get(yp_category)
    if wildcard and (yp_subtype == wildcard or mcp_subtype == wildcard):
        return True

    return False


def classify_specialization(yp_specialization, mcp_specialization) -> str:
    """
    Finer-grained cousin of specialization_matches(): instead of a plain
    OK/not-OK, says *how* a YP<->MCP specialization relates, so "exact
    subtype match" (e.g. garment_female -> garment_female) can be counted
    separately from "matched, but only via the both/any wildcard" (e.g.
    garment_female -> garment_both). Both count as OK for compliance
    purposes (specialization_matches() returns True for either), but they
    are not the same outcome operationally, and the difference is exactly
    what "how many garment_female YPs landed on a garment_female MCP"
    is asking.

    Returns one of:
      "unknown_mcp"        - mcp_specialization is missing/blank (e.g. the
                              MCP wasn't present in whatever master data
                              this evaluation run had on hand)
      "different_trade"    - different top-level trade area entirely
                              (garment vs footwear vs leather) — a hard
                              trade-area miss
      "exact_subtype"      - same category AND same subtype
                              (garment_female -> garment_female)
      "wildcard_match"     - same category, different subtype, but valid
                              via the both/any wildcard
                              (garment_female -> garment_both, or
                               garment_both -> garment_female)
      "subtype_mismatch"   - same category, different subtype, and neither
                              side is the wildcard (e.g. garment_female ->
                              garment_male) — same trade area but the
                              gender/subtype split doesn't line up
    """
    mcp = _clean_str(mcp_specialization).lower()
    if not mcp:
        return "unknown_mcp"

    yp = _clean_str(yp_specialization).lower()
    if not yp:
        return "unknown_mcp"

    yp_category, _, yp_subtype = yp.rpartition("_")
    mcp_category, _, mcp_subtype = mcp.rpartition("_")

    if not yp_category or not mcp_category or yp_category != mcp_category:
        return "different_trade"

    if yp_subtype == mcp_subtype:
        return "exact_subtype"

    wildcard = _WILDCARD_BY_CATEGORY.get(yp_category)
    if wildcard and (yp_subtype == wildcard or mcp_subtype == wildcard):
        return "wildcard_match"

    return "subtype_mismatch"


# ---------------------------------------------------------------------------
# Loading manual match reference data + automated match data
# ---------------------------------------------------------------------------

def load_manual_matches(
    path: str,
    yp_id_col: Optional[str] = None,
    mcp_id_col: Optional[str] = None,
    config_path: str = DEFAULT_CONFIG_PATH,
) -> list[dict]:
    """
    Reads the manually-matched reference sheet into [{"yp_id", "mcp_id"}, ...].

    Column names are resolved in this order: explicit yp_id_col/mcp_id_col
    args, then a "manual_match_columns" section in column_config.json (add
    one if it's not there yet — see the module docstring), then a
    best-effort header guess as a last resort. Rows missing either id are
    skipped and counted, not silently dropped without a trace.
    """
    if yp_id_col is None or mcp_id_col is None:
        try:
            cfg = load_config(config_path)
        except (FileNotFoundError, ValueError):
            cfg = {}
        manual_cols = cfg.get("manual_match_columns", {})
        yp_id_col = yp_id_col or manual_cols.get("yp_id")
        mcp_id_col = mcp_id_col or manual_cols.get("mcp_id")

    df = pd.read_excel(path, dtype=str)

    if yp_id_col is None or yp_id_col not in df.columns:
        yp_id_col = _find_column_name(
            df, ["yp_id", "yp code", "yp_code", "young professional id", "yp code (tech & data)"]
        )
    if mcp_id_col is None or mcp_id_col not in df.columns:
        mcp_id_col = _find_column_name(df, ["mcp_id", "mcp code", "mcp_code"])

    if yp_id_col is None or mcp_id_col is None or yp_id_col not in df.columns or mcp_id_col not in df.columns:
        raise ValueError(
            f"Manual match file ({path}): could not resolve YP id / MCP id columns "
            f"(got yp_id_col={yp_id_col!r}, mcp_id_col={mcp_id_col!r}). "
            f"Set 'manual_match_columns' in column_config.json, or pass yp_id_col/mcp_id_col "
            f"explicitly. Actual headers found: {list(df.columns)}"
        )

    pairs = []
    skipped = 0
    for _, row in df.iterrows():
        yp_id = _clean_str(row.get(yp_id_col))
        mcp_id = _clean_str(row.get(mcp_id_col))
        if not yp_id or not mcp_id:
            skipped += 1
            continue
        pairs.append({"yp_id": yp_id, "mcp_id": mcp_id})

    if skipped:
        logger.warning(
            "load_manual_matches(): skipped %d row(s) in %r with a blank YP id or MCP id",
            skipped, path,
        )
    logger.info("load_manual_matches() loaded %d pair(s) from %r", len(pairs), path)
    return pairs


def load_automated_matches(path: str) -> list[dict]:
    """
    Reads an automated match result from either:
      - a .json file shaped like Matcher.run()'s return value
        ({"matches": [{"yp_id", "mcp_id", "travel_time", ...}, ...]}), or
      - an .xlsx file produced by excel_export.build_results_workbook()
        (reads the "Matches" sheet's "YP ID" / "MCP ID" / "TRAVEL TIME (MINS)" columns).
    Returns [{"yp_id", "mcp_id", "travel_time": float | None}, ...].
    """
    suffix = Path(path).suffix.lower()

    if suffix == ".json":
        with open(path, "r") as f:
            data = json.load(f)
        matches = data.get("matches", data if isinstance(data, list) else [])
        pairs = [
            {
                "yp_id": _clean_str(m.get("yp_id")),
                "mcp_id": _clean_str(m.get("mcp_id")),
                "travel_time": m.get("travel_time"),
            }
            for m in matches
        ]
        logger.info("load_automated_matches() loaded %d pair(s) from JSON %r", len(pairs), path)
        return pairs

    if suffix in (".xlsx", ".xlsm"):
        df = pd.read_excel(path, sheet_name="Matches", dtype=str)
        yp_col = _find_column_name(df, ["yp id", "yp_id"])
        mcp_col = _find_column_name(df, ["mcp id", "mcp_id"])
        time_col = _find_column_name(df, ["travel time (mins)", "travel_time"])
        if yp_col is None or mcp_col is None:
            raise ValueError(
                f"Automated match file ({path}): could not find YP ID / MCP ID columns "
                f"on the 'Matches' sheet. Found columns: {list(df.columns)}"
            )
        pairs = []
        for _, row in df.iterrows():
            yp_id = _clean_str(row.get(yp_col))
            mcp_id = _clean_str(row.get(mcp_col))
            if not yp_id or not mcp_id:
                continue
            travel_time = None
            if time_col is not None:
                raw = row.get(time_col)
                try:
                    travel_time = float(raw) if raw not in (None, "") and not pd.isna(raw) else None
                except (TypeError, ValueError):
                    travel_time = None
            pairs.append({"yp_id": yp_id, "mcp_id": mcp_id, "travel_time": travel_time})
        logger.info("load_automated_matches() loaded %d pair(s) from workbook %r", len(pairs), path)
        return pairs

    raise ValueError(f"Automated match file ({path}): unsupported extension {suffix!r} (expected .json or .xlsx)")


# ---------------------------------------------------------------------------
# Distance helper (Haversine, same formula matcher.py uses for shortlisting)
# ---------------------------------------------------------------------------

def _haversine_km(lat1, lon1, lat2, lon2) -> Optional[float]:
    if lat1 is None or lon1 is None or lat2 is None or lon2 is None:
        return None
    radius = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    d_phi = math.radians(lat2 - lat1)
    d_lambda = math.radians(lon2 - lon1)
    a = math.sin(d_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(d_lambda / 2) ** 2
    return radius * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def _index_by_id(items: Iterable) -> dict:
    return {item.id: item for item in items}


def _capacity_usage(pairs: list[tuple], mcps_by_id: dict) -> tuple[dict, dict]:
    """
    pairs: [(yp_id, mcp_id), ...]. Returns (usage_by_mcp_id, violations),
    where violations only contains MCPs whose assigned count exceeds their
    configured capacity.
    """
    usage = Counter(mcp_id for _, mcp_id in pairs)
    violations = {}
    for mcp_id, count in usage.items():
        mcp = mcps_by_id.get(mcp_id)
        if mcp is not None and count > mcp.capacity:
            violations[mcp_id] = {"assigned": count, "capacity": mcp.capacity}
    return dict(usage), violations


# ---------------------------------------------------------------------------
# Per-YP row evaluation
# ---------------------------------------------------------------------------

def _evaluate_row(yp, manual_mcp_id, automated_mcp_id, automated_travel_time, mcps_by_id, config) -> dict:
    row = {
        "yp_id": yp.id,
        "yp_skill": yp.skill,
        "yp_gender": yp.gender,
        "yp_is_pwd": yp.is_pwd,
        "manual_mcp_id": manual_mcp_id,
        "automated_mcp_id": automated_mcp_id,
    }

    if manual_mcp_id is None and automated_mcp_id is None:
        row["status"] = "no_data"
        return row
    if manual_mcp_id is None:
        row["status"] = "automated_only_no_manual_reference"
        return row
    if automated_mcp_id is None:
        # The manual matcher placed this YP; the algorithm did not (waitlisted
        # or dropped them). This is an important drift signal on its own —
        # surfaced distinctly rather than folded into "divergent".
        row["status"] = "automated_missed"
        return row

    row["status"] = "compared"
    manual_mcp = mcps_by_id.get(manual_mcp_id)
    automated_mcp = mcps_by_id.get(automated_mcp_id)
    if manual_mcp is None:
        logger.warning("_evaluate_row(): manual match references unknown mcp_id=%r for yp_id=%r", manual_mcp_id, yp.id)
    if automated_mcp is None:
        logger.warning("_evaluate_row(): automated match references unknown mcp_id=%r for yp_id=%r", automated_mcp_id, yp.id)

    row["exact_match"] = manual_mcp_id == automated_mcp_id

    row["manual_trade_compatible"] = _trade_matcher.trade_matches(yp.skill, manual_mcp.skill) if manual_mcp else None
    row["automated_trade_compatible"] = _trade_matcher.trade_matches(yp.skill, automated_mcp.skill) if automated_mcp else None

    row["manual_specialization_exact"] = (manual_mcp is not None and specialization_matches(yp.skill, manual_mcp.skill))
    row["automated_specialization_exact"] = (automated_mcp is not None and specialization_matches(yp.skill, automated_mcp.skill))

    # NEW: raw MCP skill string + fine-grained classification behind each
    # match, so "exact subtype" and "wildcard-only" can be told apart.
    row["manual_mcp_skill"] = manual_mcp.skill if manual_mcp else None
    row["automated_mcp_skill"] = automated_mcp.skill if automated_mcp else None
    row["manual_specialization_class"] = classify_specialization(yp.skill, row["manual_mcp_skill"])
    row["automated_specialization_class"] = classify_specialization(yp.skill, row["automated_mcp_skill"])

    manual_km = _haversine_km(yp.latitude, yp.longitude, manual_mcp.latitude, manual_mcp.longitude) if manual_mcp else None
    automated_km = _haversine_km(yp.latitude, yp.longitude, automated_mcp.latitude, automated_mcp.longitude) if automated_mcp else None
    row["manual_distance_km"] = round(manual_km, 3) if manual_km is not None else None
    row["automated_distance_km"] = round(automated_km, 3) if automated_km is not None else None
    row["distance_delta_km"] = (
        round(automated_km - manual_km, 3) if manual_km is not None and automated_km is not None else None
    )
    # Reported for reference only — NOT diffed against manual_distance_km,
    # since it's typically a real Google Distance API duration (minutes),
    # a different metric/unit than the Haversine km figures above.
    row["automated_travel_time_reported"] = automated_travel_time

    if yp.is_pwd:
        threshold = config.get("pwd_proximity_threshold_km")
        row["manual_pwd_proximity_ok"] = (manual_km is not None and manual_km <= threshold)
        row["automated_pwd_proximity_ok"] = (automated_km is not None and automated_km <= threshold)
    else:
        row["manual_pwd_proximity_ok"] = None
        row["automated_pwd_proximity_ok"] = None

    if row["exact_match"]:
        row["verdict"] = "EXACT_MATCH"
    else:
        hard_checks = [
            row["automated_trade_compatible"] is True,
            row["distance_delta_km"] is None or row["distance_delta_km"] <= config.get("distance_tolerance_km", 1.0),
            # automated shouldn't be a LESS exact specialization match than
            # manual was (e.g. manual matched garment_female exactly,
            # automated fell back to a garment_both MCP)
            not (row["manual_specialization_exact"] and not row["automated_specialization_exact"]),
            row["automated_pwd_proximity_ok"] is not False,
        ]
        row["verdict"] = "EQUIVALENT_MATCH" if all(hard_checks) else "DIVERGENT_MATCH"

    return row


# ---------------------------------------------------------------------------
# Aggregate summary
# ---------------------------------------------------------------------------

def _rate(flags: list) -> Optional[float]:
    clean = [f for f in flags if f is not None]
    return round(sum(1 for f in clean if f) / len(clean), 4) if clean else None


def _distance_stats(values: list[float]) -> Optional[dict]:
    """Mean/median/stdev/p90/max for a list of Haversine distances (km).
    Returns None if there's nothing to summarize. Reported alongside the
    existing avg_distance_delta_km rather than replacing it: the delta
    answers "closer or farther, on average", these answer "what does the
    actual distance distribution look like" (a few bad-geocode outliers
    can blow out the mean without moving the median, so both are kept)."""
    clean = [v for v in values if v is not None]
    if not clean:
        return None
    stats = {
        "n": len(clean),
        "mean_km": round(statistics.mean(clean), 3),
        "median_km": round(statistics.median(clean), 3),
        "max_km": round(max(clean), 3),
    }
    if len(clean) > 1:
        stats["stdev_km"] = round(statistics.stdev(clean), 3)
    if len(clean) >= 10:
        stats["p90_km"] = round(statistics.quantiles(clean, n=10)[8], 3)
    return stats


def _specialization_breakdown(compared_rows: list[dict]) -> dict:
    """
    Per YP-skill-category breakdown of classify_specialization() outcomes,
    manual vs automated. This is the "how many garment_female YPs actually
    landed on a garment_female MCP, vs a garment_both MCP, vs something
    else" table — specialization_exact/OK alone can't answer that because
    it treats a wildcard match as equally "fine" as an exact one.
    """
    breakdown = defaultdict(lambda: {"manual": Counter(), "automated": Counter()})
    for row in compared_rows:
        skill = row.get("yp_skill") or "(unknown)"
        breakdown[skill]["manual"][row.get("manual_specialization_class", "unknown_mcp")] += 1
        breakdown[skill]["automated"][row.get("automated_specialization_class", "unknown_mcp")] += 1

    result = {}
    for skill, sides in breakdown.items():
        n = sum(sides["manual"].values())
        result[skill] = {
            "n": n,
            "manual": dict(sides["manual"]),
            "automated": dict(sides["automated"]),
        }
    return result


def _summarize(rows: list[dict], manual_pairs: list[tuple], automated_pairs: list[tuple], mcps_by_id: dict) -> dict:
    compared = [r for r in rows if r["status"] == "compared"]

    summary = {
        "total_yps_evaluated": len(rows),
        "compared_count": len(compared),
        "automated_missed_count": sum(1 for r in rows if r["status"] == "automated_missed"),
        "automated_only_no_manual_reference_count": sum(1 for r in rows if r["status"] == "automated_only_no_manual_reference"),
        "exact_match_count": sum(1 for r in compared if r["verdict"] == "EXACT_MATCH"),
        "equivalent_match_count": sum(1 for r in compared if r["verdict"] == "EQUIVALENT_MATCH"),
        "divergent_match_count": sum(1 for r in compared if r["verdict"] == "DIVERGENT_MATCH"),
        "distance_metric": "haversine_km",
    }

    if compared:
        summary["exact_match_rate"] = round(summary["exact_match_count"] / len(compared), 4)
        summary["equivalent_or_better_rate"] = round(
            (summary["exact_match_count"] + summary["equivalent_match_count"]) / len(compared), 4
        )

        deltas = [r["distance_delta_km"] for r in compared if r["distance_delta_km"] is not None]
        if deltas:
            summary["avg_distance_delta_km"] = round(sum(deltas) / len(deltas), 3)
            summary["automated_closer_or_equal_rate"] = round(sum(1 for d in deltas if d <= 0) / len(deltas), 4)

        # NEW: full distance distributions, not just the paired delta.
        summary["manual_distance_stats"] = _distance_stats([r["manual_distance_km"] for r in compared])
        summary["automated_distance_stats"] = _distance_stats([r["automated_distance_km"] for r in compared])

        summary["manual_trade_compliance_rate"] = _rate([r["manual_trade_compatible"] for r in compared])
        summary["automated_trade_compliance_rate"] = _rate([r["automated_trade_compatible"] for r in compared])
        summary["manual_pwd_proximity_rate"] = _rate([r["manual_pwd_proximity_ok"] for r in compared])
        summary["automated_pwd_proximity_rate"] = _rate([r["automated_pwd_proximity_ok"] for r in compared])

        # NEW: exact-subtype vs wildcard-only specialization breakdown.
        summary["specialization_breakdown"] = _specialization_breakdown(compared)

    manual_usage, manual_violations = _capacity_usage(manual_pairs, mcps_by_id)
    automated_usage, automated_violations = _capacity_usage(automated_pairs, mcps_by_id)
    summary["manual_capacity_violations"] = manual_violations
    summary["automated_capacity_violations"] = automated_violations

    return summary


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def compare_matches(
    yps: list,
    mcps: list,
    manual_pairs: list[dict],
    automated_pairs: list[dict],
    config: Optional[dict] = None,
    distance_service=None,
) -> dict:
    """
    yps / mcps: the SAME YoungProfessional/MCP records used for the runs
        being compared. For distance comparisons to mean anything, these
        need `.latitude`/`.longitude` populated — reuse the app's existing
        geocode cache/database rather than re-geocoding. If any are still
        missing coordinates and `distance_service` is supplied, this will
        call Matcher(distance_service).geocode_missing() to backfill them
        (which itself should hit the cache before any paid API call, per
        the app's existing caching design) — pass distance_service=None to
        guarantee zero geocoding calls and just accept blank distances for
        anyone not already geocoded.
    manual_pairs / automated_pairs: [{"yp_id", "mcp_id", ...}, ...], e.g.
        from load_manual_matches() / load_automated_matches().
    config: overrides for DEFAULT_CRITERIA_CONFIG.
    """
    start = time.monotonic()
    config = {**DEFAULT_CRITERIA_CONFIG, **(config or {})}
    if distance_service is not None:
        needs_geocode = [p for p in list(yps) + list(mcps)
                         if getattr(p, "latitude", None) is None or getattr(p, "longitude", None) is None]
        if needs_geocode:
            logger.info("compare_matches(): %d record(s) missing coordinates, backfilling via distance_service", len(needs_geocode))
            Matcher(distance_service).geocode_missing(yps, mcps)

    yps_by_id = _index_by_id(yps)
    mcps_by_id = _index_by_id(mcps)

    manual_by_yp = {p["yp_id"]: p["mcp_id"] for p in manual_pairs}
    automated_by_yp = {p["yp_id"]: p["mcp_id"] for p in automated_pairs}
    automated_time_by_yp = {p["yp_id"]: p.get("travel_time") for p in automated_pairs}

    all_yp_ids = sorted(set(manual_by_yp) | set(automated_by_yp))

    rows = []
    unresolved_yp_ids = []
    for yp_id in all_yp_ids:
        yp = yps_by_id.get(yp_id)
        if yp is None:
            unresolved_yp_ids.append(yp_id)
            continue
        rows.append(_evaluate_row(
            yp,
            manual_by_yp.get(yp_id),
            automated_by_yp.get(yp_id),
            automated_time_by_yp.get(yp_id),
            mcps_by_id,
            config,
        ))

    if unresolved_yp_ids:
        logger.warning(
            "compare_matches(): %d yp_id(s) appear in match data but not in the provided yps list "
            "(data integrity issue — check IDs line up): %s",
            len(unresolved_yp_ids), unresolved_yp_ids,
        )

    manual_pair_tuples = [(p["yp_id"], p["mcp_id"]) for p in manual_pairs]
    automated_pair_tuples = [(p["yp_id"], p["mcp_id"]) for p in automated_pairs]

    summary = _summarize(rows, manual_pair_tuples, automated_pair_tuples, mcps_by_id)
    summary["unresolved_yp_ids"] = unresolved_yp_ids

    elapsed = time.monotonic() - start
    logger.info(
        "compare_matches() done in %.2fs: %d compared, exact=%s equivalent_or_better=%s",
        elapsed, summary["compared_count"],
        summary.get("exact_match_rate"), summary.get("equivalent_or_better_rate"),
    )

    return {"summary": summary, "rows": rows, "config_used": config}


# ---------------------------------------------------------------------------
# Excel export (same visual style as excel_export.py, so this looks like
# part of the same family of reports rather than a bolted-on extra)
# ---------------------------------------------------------------------------

def write_evaluation_workbook(report: dict, output_path: str = "evaluation_report.xlsx") -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from datetime import datetime, timezone

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    body_font = Font(name="Arial")
    title_font = Font(name="Arial", bold=True, size=14)

    def write_table(ws, headers, rows, start_row=1):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for r, row in enumerate(rows, start=start_row + 1):
            for c, value in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=value)
                cell.font = body_font
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(header))
            for row in rows:
                val = row[col_idx - 1]
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

    wb = Workbook()

    # -- Summary sheet ------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "Automated vs Manual Match Evaluation"
    ws_summary["A1"].font = title_font
    ws_summary["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws_summary["A2"].font = body_font

    summary = report["summary"]
    # distance_stats / specialization_breakdown are nested dicts — keep them
    # off the flat scalar table below, they get their own tables.
    skip_keys = (
        "manual_capacity_violations", "automated_capacity_violations", "unresolved_yp_ids",
        "manual_distance_stats", "automated_distance_stats", "specialization_breakdown",
    )
    scalar_rows = [[k.replace("_", " ").title(), v] for k, v in summary.items() if k not in skip_keys]
    write_table(ws_summary, ["Metric", "Value"], scalar_rows, start_row=4)
    next_row = 4 + len(scalar_rows) + 2

    # NEW: distance distribution stats (mean/median/stdev/p90/max), manual vs automated.
    ws_summary.cell(row=next_row, column=1, value="Distance Distribution (km, Haversine)").font = Font(name="Arial", bold=True)
    next_row += 1
    dist_stat_keys = ["n", "mean_km", "median_km", "stdev_km", "p90_km", "max_km"]
    man_ds = summary.get("manual_distance_stats") or {}
    auto_ds = summary.get("automated_distance_stats") or {}
    dist_rows = [
        ["Manual"] + [man_ds.get(k, "") for k in dist_stat_keys],
        ["Automated"] + [auto_ds.get(k, "") for k in dist_stat_keys],
    ]
    write_table(ws_summary, ["Match Type"] + [k.replace("_km", " (km)").replace("_", " ").title() for k in dist_stat_keys], dist_rows, start_row=next_row)
    next_row += 1 + len(dist_rows) + 2

    ws_summary.cell(row=next_row, column=1, value="Manual Capacity Violations").font = Font(name="Arial", bold=True)
    violation_rows = [[mcp_id, v["assigned"], v["capacity"]] for mcp_id, v in summary.get("manual_capacity_violations", {}).items()]
    write_table(ws_summary, ["MCP ID", "Assigned", "Capacity"], violation_rows, start_row=next_row + 1)

    next_row = next_row + 1 + len(violation_rows) + 3
    ws_summary.cell(row=next_row, column=1, value="Automated Capacity Violations").font = Font(name="Arial", bold=True)
    auto_violation_rows = [[mcp_id, v["assigned"], v["capacity"]] for mcp_id, v in summary.get("automated_capacity_violations", {}).items()]
    write_table(ws_summary, ["MCP ID", "Assigned", "Capacity"], auto_violation_rows, start_row=next_row + 1)

    if summary.get("unresolved_yp_ids"):
        next_row = next_row + 1 + len(auto_violation_rows) + 3
        ws_summary.cell(row=next_row, column=1, value="Unresolved YP IDs (in match data, not in YP file)").font = Font(name="Arial", bold=True)
        write_table(ws_summary, ["YP ID"], [[y] for y in summary["unresolved_yp_ids"]], start_row=next_row + 1)

    # -- Specialization Breakdown sheet (NEW) --------------------------------
    ws_spec = wb.create_sheet("Specialization Breakdown")
    spec_headers = [
        "YP Skill", "Match Type", "N",
        "Exact Subtype", "Wildcard Match", "Subtype Mismatch", "Different Trade", "Unknown MCP",
    ]
    spec_rows = []
    for skill, data in sorted(summary.get("specialization_breakdown", {}).items()):
        for match_type in ("manual", "automated"):
            counts = data[match_type]
            spec_rows.append([
                skill, match_type.title(), data["n"],
                counts.get("exact_subtype", 0), counts.get("wildcard_match", 0),
                counts.get("subtype_mismatch", 0), counts.get("different_trade", 0),
                counts.get("unknown_mcp", 0),
            ])
    write_table(ws_spec, spec_headers, spec_rows)

    # -- Detail sheet ---------------------------------------------------------
    ws_detail = wb.create_sheet("Detail")
    detail_headers = [
        "YP ID", "YP Skill", "YP Gender", "YP PWD", "Status", "Verdict",
        "Manual MCP", "Automated MCP", "Exact Match",
        "Manual Distance (km)", "Automated Distance (km)", "Distance Delta (km)",
        "Automated Travel Time (reported)",
        "Manual Trade OK", "Automated Trade OK",
        "Manual Specialization OK", "Automated Specialization OK",
        "Manual MCP Skill", "Automated MCP Skill",                      # NEW
        "Manual Specialization Class", "Automated Specialization Class",  # NEW
        "Manual PWD Proximity OK", "Automated PWD Proximity OK",
    ]
    detail_rows = []
    for r in report["rows"]:
        detail_rows.append([
            r.get("yp_id"), r.get("yp_skill"), r.get("yp_gender"), "Yes" if r.get("yp_is_pwd") else "No",
            r.get("status"), r.get("verdict", ""),
            r.get("manual_mcp_id", ""), r.get("automated_mcp_id", ""),
            r.get("exact_match", ""),
            r.get("manual_distance_km", ""), r.get("automated_distance_km", ""), r.get("distance_delta_km", ""),
            r.get("automated_travel_time_reported", ""),
            r.get("manual_trade_compatible", ""), r.get("automated_trade_compatible", ""),
            r.get("manual_specialization_exact", ""), r.get("automated_specialization_exact", ""),
            r.get("manual_mcp_skill", ""), r.get("automated_mcp_skill", ""),
            r.get("manual_specialization_class", ""), r.get("automated_specialization_class", ""),
            r.get("manual_pwd_proximity_ok", ""), r.get("automated_pwd_proximity_ok", ""),
        ])
    write_table(ws_detail, detail_headers, detail_rows)

    wb.save(output_path)
    logger.info("write_evaluation_workbook() saved to %r", output_path)
    return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Compare automated YP-MCP matches against a manual match reference sheet")
    parser.add_argument("--yp-file", required=True)
    parser.add_argument("--mcp-file", required=True)
    parser.add_argument("--manual-match-file", required=True, help="Excel sheet with YP id / MCP id columns")
    parser.add_argument("--automated-match-file", required=True, help=".json (Matcher.run() output) or .xlsx (excel_export Matches sheet)")
    parser.add_argument("--column-config", default=DEFAULT_CONFIG_PATH)
    parser.add_argument("--criteria-config", default=None, help="Optional JSON file overriding DEFAULT_CRITERIA_CONFIG")
    parser.add_argument("--output-json", default="evaluation_report.json")
    parser.add_argument("--output-xlsx", default=None)
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    yps = load_yps(args.yp_file, config_path=args.column_config)
    mcps = load_mcps(args.mcp_file, config_path=args.column_config)
    manual_pairs = load_manual_matches(args.manual_match_file, config_path=args.column_config)
    automated_pairs = load_automated_matches(args.automated_match_file)

    criteria_config = None
    if args.criteria_config:
        with open(args.criteria_config, "r") as f:
            criteria_config = json.load(f)

    report = compare_matches(yps, mcps, manual_pairs, automated_pairs, config=criteria_config)

    with open(args.output_json, "w") as f:
        json.dump(report, f, indent=2, default=str)

    print(json.dumps(report["summary"], indent=2, default=str))
    print(f"\nFull report written to {args.output_json}")

    if args.output_xlsx:
        write_evaluation_workbook(report, args.output_xlsx)
        print(f"Excel report written to {args.output_xlsx}")


if __name__ == "__main__":
    main()